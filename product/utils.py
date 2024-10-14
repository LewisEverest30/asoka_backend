from django.db.models import Q
import openpyxl
from django.core.files import File
import re
from django.conf import settings

from .models import Gemstone, Bracelet, Stamp, Gift, GemstonePic, BraceletPic, StampPic, GiftPic


def get_cart_component_product(product_dict: dict):
    if product_dict['product_type'] == '珠':
        product = Gemstone.objects.get(id=product_dict['product_id'])
    elif product_dict['product_type'] == '手链':
        product = Bracelet.objects.get(id=product_dict['product_id'])
    elif product_dict['product_type'] == '印章':
        product = Stamp.objects.get(id=product_dict['product_id'])

    product_dict['product_name'] = product.name
    product_dict['thumbnail'] = settings.MEDIA_URL + str(product.thumbnail)
    product_dict['loc'] = product.loc
    product_dict['intro_mini'] = product.intro_mini
    product_dict['intro'] = product.intro
    product_dict['intro_full'] = product.intro_full
    product_dict['price'] = product.price



def import_gemstone_data(xls_path: str, photo_path: str):
    cover_file_default = File(open(photo_path+'/default_cover.jpg', 'rb'), name='cover.jpg')
    thumbnail_file_default = File(open(photo_path+'/default_thumbnail.png', 'rb'), name='thumbnail.jpg')
    detail_file_default = File(open(photo_path+'/default_detail.jpg', 'rb'), name='detail.jpg')
    price_default = 333
    symbol_default = '健康'
    intro_default = '没数据'
    loc_default = '没数据'
    mat_default = '没数据'
    
    workbook = openpyxl.load_workbook(xls_path)
    sheet_names = workbook.sheetnames
    sheet = workbook[sheet_names[0]]  # 选择第一个工作表

    # 遍历工作表中的数据
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 假设第一行是标题行，从第二行开始读取
        if all(cell is None for cell in row):  # 空行结束
            break

        if row[1] is not None and row[1] != '':
            mat = row[1]
        else:
            mat = mat_default

        if row[2] is not None and row[2] != '':
            loc = row[2]
        else:
            loc = loc_default

        if row[5] is not None and row[5] != '':
            price = row[5]
        else:
            price = price_default

        if row[7] is not None and row[7] != '':
            symbol = row[7].replace(',', '')
        else:
            symbol = symbol_default

        if row[9] is not None and row[9] != '':
            intro = row[9]
        else:
            intro = intro_default

        if row[8] is not None and row[8] != '':
            thumbnail = File(open(photo_path+'/thu/'+str(row[8]), 'rb'), name=str(row[8]))
        else:
            thumbnail = thumbnail_file_default
        
        if row[10] is not None and row[10] != '':
            cover = File(open(photo_path+'/1_1/'+str(row[10]), 'rb'), name=str(row[10]))
        else:
            cover = cover_file_default

        if row[11] is not None and row[11] != '':
            photo_12 = File(open(photo_path+'/1_2/'+str(row[11]), 'rb'), name=str(row[11]))
        else:
            photo_12 = detail_file_default

        # if row[12] is not None and row[12] != '':
        #     detail = File(open(photo_path+'/detail/'+str(row[12]), 'rb'), name=str(row[12]))
        # else:
        #     detail = detail_file_default
        detail = detail_file_default

        size_pattern = r'^(\d+)\D*.*'
        match_size = re.findall(size_pattern, str(row[4]))  # ['12'] 由于只有pattern一个匹配组，有多个匹配组时候是元组列表
        # print(match_size)
        size = int(match_size[0])


        new_gem = Gemstone.objects.create(  # 创建并保存模型实例
            name = row[0],
            typ = '宝石',
            symbol = symbol,
            wuxing = row[3],
            material = mat,
            position = row[6][:2],
            size = size,
            cover = cover,
            photo_12 = photo_12,
            thumbnail = thumbnail,
            detail = detail,
            loc = loc,
            price = price,
            intro = intro
        )



def import_update_gemstone_intro(xls_path: str):
    intro_default = 'test'

    workbook = openpyxl.load_workbook(xls_path)
    sheet_names = workbook.sheetnames
    sheet = workbook[sheet_names[0]]  # 选择第一个工作表

    # 遍历工作表中的数据
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 假设第一行是标题行，从第二行开始读取
        if all(cell is None for cell in row):  # 空行结束
            break
        
        gem_name = str(row[0])

        these_gem = Gemstone.objects.filter(name=gem_name)

        if row[2] is not None and row[2] != '':
            intro = row[2]
        else:
            intro = intro_default

        if row[1] is not None and row[1] != '':
            intro_15 = row[1]
        else:
            intro_15 = intro_default

        if row[3] is not None and row[3] != '':
            intro_55 = row[3]
        else:
            intro_55 = intro_default


        these_gem.update(  # 创建并保存模型实例
            intro = intro,
            intro_mini = intro_15,
            intro_full = intro_55
        )


#  bug update图片字段，路径错误
# update与create不同，create会把目标文件拷贝到media目录，并自动添加前缀
# 而update既不会拷贝也不会补充前缀
# 因此要手动给name加前缀，并手动把文件拷贝到对应目录
def import_update_gemstone_cover(xls_path: str, photo_path: str):


    workbook = openpyxl.load_workbook(xls_path)
    sheet_names = workbook.sheetnames
    sheet = workbook[sheet_names[0]]  # 选择第一个工作表

    # 遍历工作表中的数据
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 假设第一行是标题行，从第二行开始读取
        if all(cell is None for cell in row):  # 空行结束
            break
        
        gem_name = str(row[0])
        these_gem = Gemstone.objects.filter(name=gem_name)

        if row[2] is not None and row[2] != '':
            # cover = File(open(photo_path+'/'+str(row[2]), 'rb'), name=str(row[2]))
            cover = File(open(photo_path+'/'+str(row[2]), 'rb'), name='gemstone/'+str(row[2]))
        else:
            continue


        these_gem.update(  # 创建并保存模型实例
            cover = cover,
        )
    

def delete_deprecated_gemstone():

    for gem in Gemstone.objects.filter():
        cover = str(gem.cover)

        if 'cover' in cover:
            try:
                gem.delete()
            except:
                continue


def import_gemstonepics(xls_path: str, photo_path: str):


    workbook = openpyxl.load_workbook(xls_path)
    sheet_names = workbook.sheetnames
    sheet = workbook[sheet_names[0]]  # 选择第一个工作表

    # 遍历工作表中的数据
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 假设第一行是标题行，从第二行开始读取
        if all(cell is None for cell in row):  # 空行结束
            break
        
        gem_name = str(row[0])
        these_gem = Gemstone.objects.filter(name=gem_name)
        if these_gem.count() == 0:
            print(gem_name)

        for index, gem in enumerate(these_gem):

            if row[11] is not None and row[11] != '':
                pic = File(open(photo_path+'/'+str(row[11]), 'rb'), name=str(index)+str(row[11]))  # 区分文件名
            else:
                continue

            GemstonePic.objects.create(  # 创建并保存模型实例
                gemstone=gem,
                pic = pic,
            )